#!/usr/bin/env python3
"""
Verifies Salesforce translations by comparing:
  - What is in the org (from a downloaded Bilingual STF TRANSLATED section)
  - What is expected (from the master translation Excel sheet, Column C/D/E)

Generates a verification Excel with one tab per language containing:
  STF Key | Type | English Label | Translation in Org | Expected (Master) | Match
"""
import argparse
import json
import os
import sys

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


# ── Colours for Excel ────────────────────────────────────────────────────────
GREEN  = PatternFill("solid", fgColor="C6EFCE")   # ✓ Match
RED    = PatternFill("solid", fgColor="FFC7CE")   # ✗ Mismatch
YELLOW = PatternFill("solid", fgColor="FFEB9C")   # ⚠ Not in master
GREY   = PatternFill("solid", fgColor="EDEDED")   # — Missing from org
HEADER = PatternFill("solid", fgColor="1F4E79")

MATCH_LABEL   = "✓  Match"
MISS_LABEL    = "✗  Mismatch"
NO_MASTER     = "⚠  Not in Master"
MISSING_LABEL = "—  Missing from Org"


# ── Parse bilingual STF ──────────────────────────────────────────────────────

def parse_bilingual(stf_path: str, object_name: str) -> dict:
    """
    Parses the TRANSLATED section of a bilingual STF.
    Returns dict: stf_key -> {english_label, translation, out_of_date}
    Only includes keys for the specified object.
    """
    entries = {}
    in_translated = False

    with open(stf_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.rstrip("\n\r")

            if "OUTDATED AND UNTRANSLATED" in line and line.startswith("---"):
                break  # stop — we only want the TRANSLATED section

            if "TRANSLATED" in line and line.startswith("---"):
                in_translated = True
                continue

            if not in_translated or not line or line.startswith("#"):
                continue
            if "\t" not in line:
                continue

            parts = line.split("\t")
            if len(parts) < 3:
                continue

            key = parts[0].strip()
            english_label = parts[1].strip()
            translation = parts[2].strip()
            out_of_date = parts[3].strip() if len(parts) > 3 else "-"

            # Filter to the specified object
            key_parts = key.split(".")
            if len(key_parts) >= 3 and key_parts[1] == object_name:
                entries[key] = {
                    "english_label": english_label,
                    "translation": translation,
                    "out_of_date": out_of_date,
                }

    return entries


# ── Parse master sheet ───────────────────────────────────────────────────────

def load_master_sheet(master_path: str) -> dict:
    """
    Returns dict: normalized_english_label -> {spanish, portuguese}
    Col C = English label, Col D = Spanish, Col E = Portuguese.
    """
    try:
        wb = openpyxl.load_workbook(master_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"ERROR: Could not open master sheet: {e}", file=sys.stderr)
        sys.exit(1)

    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.worksheets[0]

    master = {}
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row_idx == 1:
            continue
        if not row or len(row) < 3 or not row[2]:
            continue
        english = str(row[2]).strip()
        spanish = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        portuguese = str(row[4]).strip() if len(row) > 4 and row[4] else ""
        if english and english.lower() not in master:
            master[english.lower()] = {
                "original": english,
                "spanish": spanish,
                "portuguese": portuguese,
            }

    wb.close()
    return master


# ── Classify a key ───────────────────────────────────────────────────────────

def classify_key(key: str) -> str:
    if key.startswith("CustomField.") and key.endswith(".FieldLabel"):
        return "Custom Field"
    elif key.startswith("PicklistValue."):
        return "Picklist Value"
    elif key.startswith("CustomField.") and key.endswith(".HelpText"):
        return "Help Text"
    else:
        return "Other"


# ── Compare and build rows ───────────────────────────────────────────────────

def build_verification_rows(bilingual_entries: dict, master: dict, lang_key: str) -> list:
    """
    Compares bilingual TRANSLATED entries against master sheet.
    Returns list of row dicts for the Excel output.
    """
    rows = []

    for key, entry in sorted(bilingual_entries.items()):
        english = entry["english_label"]
        org_translation = entry["translation"]
        out_of_date = entry["out_of_date"]

        master_entry = master.get(english.lower())
        if master_entry:
            expected = master_entry[lang_key]
        else:
            expected = ""

        # Determine match status
        if not expected:
            match = NO_MASTER
        elif org_translation.strip().lower() == expected.strip().lower():
            match = MATCH_LABEL
        else:
            match = MISS_LABEL

        rows.append({
            "STF Key": key,
            "Type": classify_key(key),
            "English Label": english,
            "Translation in Org": org_translation,
            "Expected (Master Sheet)": expected,
            "Match": match,
            "Out of Date": "Yes" if out_of_date == "*" else "No",
        })

    return rows


# ── Write Excel ──────────────────────────────────────────────────────────────

def write_verification_excel(output_path: str, tabs: dict):
    """
    tabs: dict of sheet_name -> list of row dicts
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    col_widths = {
        "STF Key": 55,
        "Type": 15,
        "English Label": 35,
        "Translation in Org": 35,
        "Expected (Master Sheet)": 35,
        "Match": 20,
        "Out of Date": 12,
    }

    match_fills = {
        MATCH_LABEL:   GREEN,
        MISS_LABEL:    RED,
        NO_MASTER:     YELLOW,
        MISSING_LABEL: GREY,
    }

    for sheet_name, rows in tabs.items():
        ws = wb.create_sheet(sheet_name)
        headers = list(col_widths.keys())

        # Header row
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = HEADER
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(wrap_text=False)
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[header]

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        # Data rows
        for row_idx, row in enumerate(rows, 2):
            match_val = row.get("Match", "")
            fill = match_fills.get(match_val)
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))
                if fill:
                    cell.fill = fill

    wb.save(output_path)


# ── Summary stats ────────────────────────────────────────────────────────────

def summarise(rows: list, lang: str):
    counts = {MATCH_LABEL: 0, MISS_LABEL: 0, NO_MASTER: 0, MISSING_LABEL: 0}
    for r in rows:
        counts[r["Match"]] = counts.get(r["Match"], 0) + 1
    total = len(rows)
    print(f"\n  {lang}:")
    print(f"    Total entries in org:   {total}")
    print(f"    {MATCH_LABEL}:          {counts[MATCH_LABEL]}")
    print(f"    {MISS_LABEL}:       {counts[MISS_LABEL]}")
    print(f"    {NO_MASTER}:    {counts[NO_MASTER]}")


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Verify Salesforce translations against master sheet")
    parser.add_argument("--object",       required=True, dest="object_name")
    parser.add_argument("--master",       required=True)
    parser.add_argument("--output",       required=True)
    parser.add_argument("--bilingual-es", default="")
    parser.add_argument("--bilingual-pt", default="")
    args = parser.parse_args()

    if not args.bilingual_es and not args.bilingual_pt:
        print("ERROR: Provide at least one bilingual STF (--bilingual-es or --bilingual-pt)", file=sys.stderr)
        sys.exit(1)

    print(f"Loading master sheet...")
    master = load_master_sheet(args.master)
    print(f"  {len(master)} unique English labels loaded")

    os.makedirs(args.output, exist_ok=True)
    tabs = {}

    for lang_label, stf_path, lang_key in [
        ("Spanish (es)",          args.bilingual_es, "spanish"),
        ("Portuguese (pt_BR)",    args.bilingual_pt, "portuguese"),
    ]:
        if not stf_path:
            continue
        if not os.path.isfile(stf_path):
            print(f"WARNING: File not found: {stf_path}", file=sys.stderr)
            continue

        print(f"Parsing bilingual STF: {os.path.basename(stf_path)}...")
        entries = parse_bilingual(stf_path, args.object_name)
        print(f"  {len(entries)} translated entries found for {args.object_name}")

        rows = build_verification_rows(entries, master, lang_key)
        tabs[lang_label] = rows
        summarise(rows, lang_label)

    if not tabs:
        print("ERROR: No valid bilingual STF files could be parsed.", file=sys.stderr)
        sys.exit(1)

    output_file = os.path.join(args.output, f"{args.object_name}_verification.xlsx")
    write_verification_excel(output_file, tabs)

    print(f"\nVerification Excel written to: {output_file}")
    print(json.dumps({"status": "ok", "output": output_file}, indent=2))


if __name__ == "__main__":
    main()
