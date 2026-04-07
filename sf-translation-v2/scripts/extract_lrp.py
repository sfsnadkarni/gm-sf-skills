#!/usr/bin/env python3
"""
Extracts Lightning Record Page (LRP) tab and related list labels for translation.

Parses the flexipage XML to find all tab and related list label references,
queries the org for existing custom label translations, and matches against
the master translation sheet.

Outputs a JSON file consumed by generate_labels_stf.py.

Usage:
  python3 extract_lrp.py \\
    --flexipage path/to/Object_Record_Page.flexipage-meta.xml \\
    --org my-org-alias \\
    --master path/to/Master_Sheet.xlsx \\
    --object Vehicle \\
    --output ~/Desktop/sf-translation-output \\
    [--existing-es path/to/bilingual_es.stf] \\
    [--existing-pt path/to/bilingual_pt_BR.stf]
"""
import argparse
import json
import os
import re
import subprocess
import sys
import xml.etree.ElementTree as ET

import openpyxl


MNS = "http://soap.sforce.com/2006/04/metadata"


def _t(tag: str) -> str:
    return f"{{{MNS}}}{tag}"


def _derive_api_name(raw_title: str) -> str:
    """Derive a Custom Label API name from a plain-text tab title."""
    name = re.sub(r"[^a-zA-Z0-9_ ]", "", raw_title).strip()
    name = re.sub(r"\s+", "_", name)
    return name[:40] or "Custom_Label"


def parse_lrp(flexipage_path: str) -> list:
    """
    Parse flexipage XML and return a list of component dicts:
      {component_type, component_name, raw_title, label_type, label_api_name}

    label_type values:
      "custom_label"  — value was {!$Label.ApiName}  → query org
      "standard_tab"  — value was Standard.Tab.xxx   → skip
      "plain_text"    — literal string                → look up in master sheet
      "empty"         — no title/label found          → skip
    """
    tree = ET.parse(flexipage_path)
    root = tree.getroot()
    components = []

    def get_prop(instance_el, prop_name):
        for prop in instance_el.findall(f".//{_t('componentInstanceProperties')}"):
            name_el  = prop.find(_t("name"))
            value_el = prop.find(_t("value"))
            if name_el is not None and name_el.text == prop_name:
                return value_el.text if value_el is not None else ""
        return None

    def classify(raw_value):
        if not raw_value:
            return "empty", None
        m = re.match(r"^\{!\$Label\.(\w+)\}$", raw_value.strip())
        if m:
            return "custom_label", m.group(1)
        if raw_value.startswith("Standard.Tab."):
            return "standard_tab", None
        return "plain_text", None

    for region in root.findall(_t("flexiPageRegions")):
        for item in region.findall(_t("itemInstances")):
            comp = item.find(_t("componentInstance"))
            if comp is None:
                continue
            comp_name_el = comp.find(_t("componentName"))
            if comp_name_el is None:
                continue
            comp_name = comp_name_el.text or ""

            # Tabs
            if comp_name == "flexipage:tab":
                title = get_prop(comp, "title")
                label_type, label_api = classify(title)
                components.append({
                    "component_type": "Tab",
                    "component_name": comp_name,
                    "raw_title":      title or "",
                    "label_type":     label_type,
                    "label_api_name": label_api,
                    "derived_api_name": _derive_api_name(title or "") if label_type == "plain_text" else None,
                })

            # Related lists — only the title uses a custom label;
            # individual fields are handled by regular CustomField STF entries (v1).
            elif comp_name in (
                "lst:dynamicRelatedList",
                "forceApps:relatedListContainer",
                "flexipage:relatedListSingle",
                "force:relatedList",
                "flexipage:relatedList",
            ):
                rl_label = (
                    get_prop(comp, "relatedListLabel")
                    or get_prop(comp, "title")
                    or get_prop(comp, "label")
                )
                rl_api = get_prop(comp, "relatedListApiName") or ""
                label_type, label_api = classify(rl_label)
                components.append({
                    "component_type":    "RelatedList",
                    "component_name":    rl_api or comp_name,
                    "raw_title":         rl_label or "",
                    "label_type":        label_type,
                    "label_api_name":    label_api,
                    "derived_api_name":  _derive_api_name(rl_label or "") if label_type == "plain_text" else None,
                })

    return components


def load_master(master_path: str) -> dict:
    """Return {normalized_english: {english, spanish, portuguese}} from Sheet1 Col C/D/E."""
    wb = openpyxl.load_workbook(master_path, read_only=True, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.worksheets[0]
    master = {}
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            continue
        if not row or len(row) < 3:
            continue
        eng = str(row[2]).strip() if row[2] else ""
        es  = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        pt  = str(row[4]).strip() if len(row) > 4 and row[4] else ""
        if eng:
            master[eng.lower()] = {"english": eng, "spanish": es, "portuguese": pt}
    wb.close()
    return master


def load_existing_label_keys(stf_path: str) -> set:
    """Return set of customLabel.* keys already present in an existing STF."""
    keys = set()
    if not stf_path or not os.path.isfile(stf_path):
        return keys
    is_bilingual = False
    in_translated = False
    with open(stf_path, encoding="utf-8") as f:
        for line in f:
            line = line.rstrip("\n\r")
            if line.startswith("Type:"):
                ftype = line.split(":", 1)[1].strip().lower()
                is_bilingual = (ftype == "bilingual")
                if not is_bilingual:
                    in_translated = True
                continue
            if is_bilingual:
                if line.startswith("---") and "TRANSLATED" in line and "UNTRANSLATED" not in line:
                    in_translated = True
                    continue
                if line.startswith("---") and "UNTRANSLATED" in line:
                    in_translated = False
                    continue
            if not in_translated or not line or line.startswith("#") or "\t" not in line:
                continue
            key = line.split("\t", 1)[0].strip()
            if key.startswith("customLabel."):
                keys.add(key)
    return keys


def sf_tooling_query(soql: str, org: str) -> list:
    """Run a Tooling API query and return records list."""
    cmd = ["sf", "data", "query", "--json", "--use-tooling-api", "-q", soql, "-o", org]
    try:
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
        data = json.loads(r.stdout or "{}")
        if data.get("status") != 0:
            print(f"  [warn] Query error: {data.get('message', '')} | SOQL: {soql[:120]}",
                  file=sys.stderr)
        return data.get("result", {}).get("records", [])
    except Exception as e:
        print(f"  [warn] Query failed: {e} | SOQL: {soql[:120]}", file=sys.stderr)
        return []


def query_label_english_values(label_api_names: list, org: str) -> dict:
    """Return {label_api_name: english_value} for all given label names."""
    if not label_api_names:
        return {}
    names_quoted = ", ".join(f"'{n}'" for n in label_api_names)
    soql = f"SELECT Name, Value FROM ExternalString WHERE Name IN ({names_quoted})"
    print(f"  Querying org for English values of {len(label_api_names)} label(s)...")
    records = sf_tooling_query(soql, org)
    return {rec["Name"]: rec.get("Value", "") for rec in records if rec.get("Name")}


def query_label_existing_translations(label_api_names: list, org: str) -> dict:
    """
    Return {label_api_name: {es: str, pt_BR: str}} for existing translations in org.
    Two-step: get IDs from ExternalString, then query ExternalStringLocalization.
    """
    if not label_api_names:
        return {}

    names_quoted = ", ".join(f"'{n}'" for n in label_api_names)
    id_soql = f"SELECT Id, Name FROM ExternalString WHERE Name IN ({names_quoted})"
    id_records = sf_tooling_query(id_soql, org)
    if not id_records:
        return {}

    id_to_name = {rec["Id"]: rec["Name"] for rec in id_records}
    ids_quoted  = ", ".join(f"'{i}'" for i in id_to_name)

    trans_soql = (
        f"SELECT ExternalStringId, Language, Value FROM ExternalStringLocalization "
        f"WHERE ExternalStringId IN ({ids_quoted}) AND Language IN ('es', 'pt_BR')"
    )
    print(f"  Querying org for existing translations of {len(label_api_names)} label(s)...")
    trans_records = sf_tooling_query(trans_soql, org)

    result = {}
    for rec in trans_records:
        name = id_to_name.get(rec.get("ExternalStringId", ""))
        if not name:
            continue
        result.setdefault(name, {})
        lang = rec.get("Language", "")
        val  = rec.get("Value", "")
        if lang == "es":
            result[name]["es"] = val
        elif lang == "pt_BR":
            result[name]["pt_BR"] = val

    return result


def is_multi_value(value: str) -> bool:
    return bool(value) and value.count(",") > 1


def main():
    parser = argparse.ArgumentParser(description="Extract LRP label data for translation")
    parser.add_argument("--flexipage", required=True, help="Path to flexipage XML file")
    parser.add_argument("--org",       required=True, help="Salesforce org alias or username")
    parser.add_argument("--master",    required=True, help="Path to master translation Excel")
    parser.add_argument("--object",    required=True, dest="object_name", help="Salesforce object API name")
    parser.add_argument("--output",    required=True, help="Output directory")
    parser.add_argument("--existing-es", default="", help="Existing Spanish custom label STF (optional)")
    parser.add_argument("--existing-pt", default="", help="Existing Portuguese custom label STF (optional)")
    args = parser.parse_args()

    os.makedirs(args.output, exist_ok=True)

    # ── Parse LRP ─────────────────────────────────────────────────────────────
    print(f"Parsing flexipage: {os.path.basename(args.flexipage)}")
    try:
        components = parse_lrp(args.flexipage)
    except Exception as e:
        print(f"ERROR: Could not parse flexipage: {e}", file=sys.stderr)
        sys.exit(1)

    tabs      = [c for c in components if c["component_type"] == "Tab"]
    rel_lists = [c for c in components if c["component_type"] == "RelatedList"]
    print(f"  Found {len(tabs)} Tab(s), {len(rel_lists)} Related List(s)")

    custom_label_refs = [c for c in components if c["label_type"] == "custom_label"]
    plain_text_comps  = [c for c in components if c["label_type"] == "plain_text"]
    standard_tabs     = [c for c in components if c["label_type"] == "standard_tab"]
    label_api_names   = list({c["label_api_name"] for c in custom_label_refs if c["label_api_name"]})

    # ── Load existing STF keys ────────────────────────────────────────────────
    existing_es_keys = load_existing_label_keys(args.existing_es)
    existing_pt_keys = load_existing_label_keys(args.existing_pt)
    print(f"  Existing label keys: {len(existing_es_keys)} ES, {len(existing_pt_keys)} PT")

    # ── Load master sheet ─────────────────────────────────────────────────────
    print(f"Loading master sheet: {os.path.basename(args.master)}")
    master = load_master(args.master)
    print(f"  {len(master)} unique English labels")

    # ── Query org ─────────────────────────────────────────────────────────────
    org_translations   = query_label_existing_translations(label_api_names, args.org)
    label_english_vals = query_label_english_values(label_api_names, args.org)

    # ── Classify $Label refs ──────────────────────────────────────────────────
    label_already_translated = []
    label_needs_translation  = []
    label_not_in_org         = []

    for c in custom_label_refs:
        api = c["label_api_name"]
        stf_key   = f"customLabel.{api}"
        org_trans = org_translations.get(api, {})
        has_es    = bool(org_trans.get("es"))    or (stf_key in existing_es_keys)
        has_pt    = bool(org_trans.get("pt_BR")) or (stf_key in existing_pt_keys)

        if has_es and has_pt:
            label_already_translated.append({
                **c,
                "org_es": org_trans.get("es"),
                "org_pt": org_trans.get("pt_BR"),
            })
            continue

        eng_value = label_english_vals.get(api, "")
        if not eng_value:
            label_not_in_org.append(c)
            continue

        master_entry = master.get(eng_value.lower())
        if master_entry:
            es_val       = master_entry["spanish"].strip()
            pt_val       = master_entry["portuguese"].strip()
            is_multi_es  = is_multi_value(es_val)
            is_multi_pt  = is_multi_value(pt_val)
            label_needs_translation.append({
                **c,
                "english_value": eng_value,
                "spanish":       es_val,
                "portuguese":    pt_val,
                "has_es":        has_es,
                "has_pt":        has_pt,
                "write_es":      not has_es and bool(es_val) and not is_multi_es,
                "write_pt":      not has_pt and bool(pt_val) and not is_multi_pt,
                "miss_es":       not has_es and (not es_val or is_multi_es),
                "miss_pt":       not has_pt and (not pt_val or is_multi_pt),
                "in_master":     True,
            })
        else:
            label_needs_translation.append({
                **c,
                "english_value": eng_value,
                "spanish":       "",
                "portuguese":    "",
                "has_es":        has_es,
                "has_pt":        has_pt,
                "write_es":      False,
                "write_pt":      False,
                "miss_es":       not has_es,
                "miss_pt":       not has_pt,
                "in_master":     False,
            })

    # ── Classify plain text ───────────────────────────────────────────────────
    plain_matched        = []
    plain_unmatched      = []
    plain_skipped_brand  = []

    for c in plain_text_comps:
        key = c["raw_title"].lower()
        if key in master:
            entry = master[key]
            es_same = (entry["spanish"].strip().lower() == c["raw_title"].lower())
            pt_same = (entry["portuguese"].strip().lower() == c["raw_title"].lower())
            if es_same and pt_same:
                plain_skipped_brand.append({**c, **entry})
            else:
                es_val      = entry["spanish"].strip()
                pt_val      = entry["portuguese"].strip()
                plain_matched.append({
                    **c,
                    "english":    entry["english"],
                    "spanish":    es_val,
                    "portuguese": pt_val,
                    "write_es":   bool(es_val) and not is_multi_value(es_val),
                    "write_pt":   bool(pt_val) and not is_multi_value(pt_val),
                    "miss_es":    not es_val or is_multi_value(es_val),
                    "miss_pt":    not pt_val or is_multi_value(pt_val),
                })
        else:
            plain_unmatched.append(c)

    # ── Build output ──────────────────────────────────────────────────────────
    need_in_master  = [c for c in label_needs_translation if c.get("in_master")]
    need_not_master = [c for c in label_needs_translation if not c.get("in_master")]

    # Consolidate miss report: remove per-language rows for labels not in master,
    # replacing with a single combined row.
    # (handled in generate_labels_stf.py)

    stats = {
        "tabs_total":          len(tabs),
        "related_lists_total": len(rel_lists),
        "standard_tabs_skipped": len(standard_tabs),
        "brand_names_skipped": len(plain_skipped_brand),
        "already_translated":  len(label_already_translated),
        "labels_write_es":     sum(1 for c in label_needs_translation if c.get("write_es")),
        "labels_write_pt":     sum(1 for c in label_needs_translation if c.get("write_pt")),
        "plain_new_labels_write_es": sum(1 for c in plain_matched if c.get("write_es")),
        "plain_new_labels_write_pt": sum(1 for c in plain_matched if c.get("write_pt")),
        "labels_not_in_org":   len(label_not_in_org),
        "labels_not_in_master": len(need_not_master),
        "plain_unmatched":     len(plain_unmatched),
    }

    output_data = {
        "object":                  args.object_name,
        "flexipage":               args.flexipage,
        "label_already_translated": label_already_translated,
        "label_needs_translation": label_needs_translation,
        "plain_matched":           plain_matched,
        "label_not_in_org":        label_not_in_org,
        "plain_unmatched":         plain_unmatched,
        "plain_skipped_brand":     plain_skipped_brand,
        "standard_tabs_skipped":   len(standard_tabs),
        "stats":                   stats,
    }

    output_path = os.path.join(args.output, f"{args.object_name}_lrp_matches.json")
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output_data, f, indent=2, ensure_ascii=False)

    total_write_es = stats["labels_write_es"] + stats["plain_new_labels_write_es"]
    total_write_pt = stats["labels_write_pt"] + stats["plain_new_labels_write_pt"]

    print(json.dumps({
        "status":         "ok",
        "lrp_matches":    output_path,
        "summary": (
            f"Tabs: {len(tabs)}, Related lists: {len(rel_lists)}. "
            f"Standard/brand skipped: {len(standard_tabs) + len(plain_skipped_brand)}. "
            f"ES entries to write: {total_write_es}. "
            f"PT entries to write: {total_write_pt}. "
            f"New custom labels to create: {len(plain_matched)}."
        ),
        "stats": stats,
    }, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
