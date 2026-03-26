#!/usr/bin/env python3
"""
Compares labels in the intermediate Excel against Column C of the Master Sheet.

Master Sheet (Sheet1) layout:
  Column A: Object Name   (ignored — matching is label-only)
  Column B: Field Type    (ignored)
  Column C: Field Name    (English label — matched against)
  Column D: Spanish Translation
  Column E: Portuguese Translation
  Row 1 is the header.

Intermediate Excel layout:
  Tab Custom_Fields:    Row 1 = header | Field Label | Field API Name
  Tab Picklist_Values:  Row 1 = header | Field API Name | Picklist Value | Picklist Label

Outputs a JSON file with full match data for downstream scripts.
"""
import argparse
import json
import os
import sys

import openpyxl


def is_multi_value(value: str) -> bool:
    """
    Detects if a translation cell contains multiple comma-separated values.
    A single comma within a normal phrase (e.g. "Fleet Account Number (FAN)") is
    acceptable, but multiple commas or commas between repeated phrases indicate
    a multi-value cell that needs manual review.
    """
    if not value:
        return False
    # If the cell contains more than one comma, treat it as multi-value
    return value.count(",") > 1


def load_master_sheet(master_path: str) -> dict:
    """
    Returns a dict: normalized_english_label -> {spanish, portuguese, multi_value_es, multi_value_pt}
    Reads Sheet1. Row 1 is header. Col C=label, Col D=spanish, Col E=portuguese.
    """
    try:
        wb = openpyxl.load_workbook(master_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"ERROR: Could not open master sheet: {e}", file=sys.stderr)
        sys.exit(1)

    if "Sheet1" not in wb.sheetnames:
        # Fall back to first sheet
        ws = wb.worksheets[0]
        print(f"WARNING: 'Sheet1' not found; using first sheet '{ws.title}'", file=sys.stderr)
    else:
        ws = wb["Sheet1"]

    translations = {}  # normalized_label -> {spanish, portuguese, original_label}

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            continue  # skip header
        if not row or len(row) < 3:
            continue

        english_label = row[2]  # Column C (0-indexed)
        spanish = row[3] if len(row) > 3 else None   # Column D
        portuguese = row[4] if len(row) > 4 else None  # Column E

        if not english_label:
            continue

        english_str = str(english_label).strip()
        if not english_str:
            continue

        es_str = str(spanish).strip() if spanish else ""
        pt_str = str(portuguese).strip() if portuguese else ""

        normalized = english_str.lower()
        # Keep first occurrence if duplicates exist
        if normalized not in translations:
            translations[normalized] = {
                "original_label": english_str,
                "spanish": es_str,
                "portuguese": pt_str,
                "multi_value_es": is_multi_value(es_str),
                "multi_value_pt": is_multi_value(pt_str),
            }

    wb.close()
    return translations


def load_intermediate_excel(intermediate_path: str):
    """
    Returns:
      fields: list of {field_label, field_api_name}
      picklist_values: list of {field_api_name_stf, picklist_value, picklist_label}
    """
    try:
        wb = openpyxl.load_workbook(intermediate_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"ERROR: Could not open intermediate Excel: {e}", file=sys.stderr)
        sys.exit(1)

    fields = []
    picklist_values = []

    # Custom_Fields tab
    if "Custom_Fields" in wb.sheetnames:
        ws = wb["Custom_Fields"]
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx == 1:
                continue  # skip header
            if not row or len(row) < 2:
                continue
            label = str(row[0]).strip() if row[0] else ""
            api_name = str(row[1]).strip() if row[1] else ""
            if label and api_name:
                # Derive STF field name (strip __c)
                stf_name = api_name[:-3] if api_name.endswith("__c") else api_name
                fields.append({
                    "field_label": label,
                    "field_api_name": api_name,
                    "stf_field_name": stf_name,
                })

    # Picklist_Values tab
    if "Picklist_Values" in wb.sheetnames:
        ws = wb["Picklist_Values"]
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx == 1:
                continue  # skip header
            if not row or len(row) < 3:
                continue
            stf_field_name = str(row[0]).strip() if row[0] else ""
            pv_value = str(row[1]).strip() if row[1] else ""
            pv_label = str(row[2]).strip() if row[2] else ""
            if stf_field_name and pv_value:
                picklist_values.append({
                    "stf_field_name": stf_field_name,
                    "picklist_value": pv_value,
                    "picklist_label": pv_label or pv_value,
                })

    wb.close()
    return fields, picklist_values


def match_against_master(fields: list, picklist_values: list, master: dict):
    """
    Matches field labels and picklist labels against master sheet Column C.

    Returns:
      matched_fields: list of {field_label, stf_field_name, spanish, portuguese}
      unmatched_fields: list of {field_label, stf_field_name}
      matched_picklists: list of {stf_field_name, picklist_value, picklist_label, spanish, portuguese}
      unmatched_picklists: list of {stf_field_name, picklist_value, picklist_label}
    """
    matched_fields = []
    unmatched_fields = []

    for f in fields:
        key = f["field_label"].lower()
        if key in master:
            entry = master[key]
            matched_fields.append({
                "field_label": f["field_label"],
                "stf_field_name": f["stf_field_name"],
                "field_api_name": f["field_api_name"],
                "spanish": entry["spanish"],
                "portuguese": entry["portuguese"],
                "multi_value_es": entry.get("multi_value_es", False),
                "multi_value_pt": entry.get("multi_value_pt", False),
            })
        else:
            unmatched_fields.append({
                "field_label": f["field_label"],
                "stf_field_name": f["stf_field_name"],
                "field_api_name": f["field_api_name"],
            })

    matched_picklists = []
    unmatched_picklists = []

    for pv in picklist_values:
        key = pv["picklist_label"].lower()
        if key in master:
            entry = master[key]
            matched_picklists.append({
                "stf_field_name": pv["stf_field_name"],
                "picklist_value": pv["picklist_value"],
                "picklist_label": pv["picklist_label"],
                "spanish": entry["spanish"],
                "portuguese": entry["portuguese"],
                "multi_value_es": entry.get("multi_value_es", False),
                "multi_value_pt": entry.get("multi_value_pt", False),
            })
        else:
            unmatched_picklists.append({
                "stf_field_name": pv["stf_field_name"],
                "picklist_value": pv["picklist_value"],
                "picklist_label": pv["picklist_label"],
            })

    return matched_fields, unmatched_fields, matched_picklists, unmatched_picklists


def main():
    parser = argparse.ArgumentParser(description="Match intermediate Excel labels against master sheet")
    parser.add_argument("--intermediate", required=True, help="Path to intermediate Excel file")
    parser.add_argument("--master", required=True, help="Path to master translation Excel")
    parser.add_argument("--output", required=True, help="Output path for matches JSON file")
    args = parser.parse_args()

    print("Loading master translation sheet...")
    master = load_master_sheet(args.master)
    print(f"  Master sheet loaded: {len(master)} unique English labels")

    print("Loading intermediate Excel...")
    fields, picklist_values = load_intermediate_excel(args.intermediate)
    print(f"  Intermediate: {len(fields)} custom field(s), {len(picklist_values)} picklist value(s)")

    print("Matching labels...")
    matched_f, unmatched_f, matched_pv, unmatched_pv = match_against_master(fields, picklist_values, master)

    result = {
        "matched_fields": matched_f,
        "unmatched_fields": unmatched_f,
        "matched_picklists": matched_pv,
        "unmatched_picklists": unmatched_pv,
        "stats": {
            "fields_total": len(fields),
            "fields_matched": len(matched_f),
            "fields_unmatched": len(unmatched_f),
            "picklists_total": len(picklist_values),
            "picklists_matched": len(matched_pv),
            "picklists_unmatched": len(unmatched_pv),
        },
    }

    os.makedirs(os.path.dirname(args.output) if os.path.dirname(args.output) else ".", exist_ok=True)
    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, ensure_ascii=False)

    stats = result["stats"]
    print(json.dumps({
        "status": "ok",
        "matches_file": args.output,
        "summary": (
            f"Fields: {stats['fields_matched']}/{stats['fields_total']} matched. "
            f"Picklist values: {stats['picklists_matched']}/{stats['picklists_total']} matched."
        ),
        "stats": stats,
    }, indent=2))


if __name__ == "__main__":
    main()
